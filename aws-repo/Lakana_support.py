'''
Check list:

1. Better practice to store API call output in a variable.
2. Avoid calling API calls in loops. Instead use variables.

Documentations:

Graphite -- http://graphite-api.readthedocs.io/en/latest/
            http://graphite.readthedocs.io/en/latest/
Grafana -- https://www.hostedgraphite.com/docs/advanced/grafana-api.html
Docker -- https://docker-py.readthedocs.io/en/stable/
AWS -- https://boto3.readthedocs.io/en/latest/
Varnish -- https://varnish-cache.org/
django -- https://docs.djangoproject.com/en/2.0/

Useful links:
http://ipython.readthedocs.io/en/stable/interactive/magics.html
https://www.dataquest.io/blog/learn-data-science/
http://docs.activestate.com/activepython/3.3/pywin32/html/com/win32com/HTML/QuickStartClientCom.html
http://timgolden.me.uk/python/win32_how_do_i/watch_directory_for_changes.html
https://www.digitalocean.com/community/tutorials/how-to-use-web-apis-in-python-3
https://www.digitalocean.com/community/tutorials/how-to-scrape-web-pages-with-beautiful-soup-and-python-3
https://www.digitalocean.com/community/tutorials/how-to-work-with-web-data-using-requests-and-beautiful-soup-with-python-3

'''

# python modules that you should aware of to do basic automations

import requests
import win32com.client
import pywinauto #https://pywinauto.readthedocs.io/en/latest/getting_started.html
                 #https://media.readthedocs.org/pdf/pywinauto/latest/pywinauto.pdf
import pyautogui
import urllib.request  #Equivalent to urllib2 in python3
import logging
import boto3
import subprocess
import json
import random
import os
import sys
import fabric3 # http://docs.fabfile.org/en/1.14/tutorial.html
import re
import glob
import fnmatch
import selenium
import django
import xlwings
import openpyxl # https://openpyxl.readthedocs.io/en/stable/tutorial.html
import xlsxwriter # http://xlsxwriter.readthedocs.io/index.html
import pynput # https://pypi.python.org/pypi/pynput
import getpass # https://docs.python.org/3/library/getpass.html
import pyHook # http://pyhook.sourceforge.net/doc_1.5.0/
import pythoncom
import pwd # https://docs.python.org/3/library/pwd.html#module-pwd
import pandas
import numpy

#Setting up assumerole on boto3

sts=boto3.client('sts')
res=sts.assume_role(RoleArn='arn:aws:iam::462397709356:role/lakana-fox-admin',RoleSessionName='testrole')

s3=boto3.client('s3',aws_access_key_id=res['Credentials']['AccessKeyId'],
aws_secret_access_key=res['Credentials']['SecretAccessKey'],
aws_session_token=res['Credentials']['SessionToken'])  


#listing buckets and objects of each bucket

for i in range(len(s3.list_buckets()['Buckets'])):
    buck=s3.list_buckets()['Buckets'][i]['Name']
    pprint.pprint(s3.list_objects(Bucket=buck))
    time.sleep(1000)
    
#Listing each object that CacheControl metadata set

API_list_buckets
count=0
for i in range(len(s3.list_buckets()['Buckets'])):
    buck=API_list_buckets['Buckets'][i]['Name']
    pprint.pprint('Bucket=',buck)
    print('\n'*3)
    if count==4:
        break
    try:
        for j in range(len(s3.list_objects(Bucket=buck)['Contents'])):
            obj=s3.list_objects(Bucket=buck)['Contents'][j]['Key']
#            pprint.pprint("Object=",obj)
            rep=s3.get_object(Bucket=buck,Key=obj)
            if rep['CacheControl']:
                print(obj,"set CacheControl=",rep['CacheControl'])
                count=count+1
                
    except:
        pass


#listing buckets, objects, Keys 

s3l=boto3.client('s3') #Creating s3 boto3 object
s3_buckets=s3l.list_buckets() #Listing buckets under the account

s3_objects=s3l.list_objects(Bucket='lakana-cloudformation')['Contents']
no_of_objs=len(s3_objects)
for i in range(no_of_objs):
    print(s3_obje_ts[i]['Key']) #Listing object keys under specific bucket

s3_objects=s3l.list_objects(Bucket='lakana-cloudformation')['Contents']
no_of_objs=len(_)
for i in range(no_of_objs):
    if s3_objects[i]['Key'].startswith('platform'):
        print(s3_objects[i]['Key']) #Listing object keys those starts with specific word

'''
Exam
import boto3
s3 = boto3.resource('s3')
s3.meta.client.download_file('mybucket', 'hello.txt', '/tmp/hello.txt')

'''

#Downloading objects from specific s3 bucket with specific word they start with

''' https://www.peterbe.com/plog/fastest-way-to-download-a-file-from-s3 
    https://alexwlchan.net/2017/07/listing-s3-keys/
'''

s3_objects=s3l.list_objects(Bucket='lakana-cloudformation')['Contents']
no_of_objs=len(s3_objects)
for i in range(no_of_objs):
    if s3_objects[i]['Key'].startswith('platform/v37'):
        obj_download=s3_objects[i]['Key']
        s3l.download_file('lakana-cloudformation',obj_download,obj_download.split('/')[-1])



#Deleting security groups from the account


ec2=boto3.client('ec2')

sg=ec2.describe_security_groups()

#print("List of security groups:")
for i in range(len(sg['SecurityGroups'])):
    if sg['SecurityGroups'][i]['GroupName'] == 'default':
        print("Default SG can't be deleted")
    else:
        try:
            del_g=ec2.delete_security_group(GroupName=sg['SecurityGroups'][i]['GroupName'])
            if del_g['ResponseMetadata']['HTTPStatusCode']==200:
                print(sg['SecurityGroups'][i]['GroupName'],"is deleted")
        except:
            print("Some error occurred while deleting ",sg['SecurityGroups'][i]['GroupName'])


#Remove directory from git and local
'''
You could checkout 'master' with both directories;

git rm -r one-of-the-directories
git commit -m "Remove duplicated directory"
git push origin <your-git-branch> (typically 'master', but not always)
Remove directory from git but NOT local
As mentioned in the comments, what you usually want to do is remove this directory from git but not delete it entirely from the filesystem (local)

In that case use:

git rm -r --cached myFolder

'''            

#Solution for json serialize a datetime object and dumping dict object into an external json file.

import datetime
import json

def myconverter(o):
    if isinstance(o, datetime.datetime):
        return o.__str__()

with open('data.txt', 'w') as outfile:
    json.dump(data, outfile, default=myconverter) #json.dumps(data,default=myconverter)

#Listing instances Names, instance types

import boto3
import openpyxl as el
ec2=boto3.client("ec2")

data=ec2.describe_instances()   

''' goto with #8 '''

count=0
wb=el.Workbook()
sh1=wb.create_sheet("Avg CPU")
sh1['A1']="Instace Name"
sh1['B1']="Instance Type"
sh1['C1']="State"
sh1['D1']="Avg CPU"

for i in range(len(data['Reservations'])):
    for j in range(len(data['Reservations'][i]['Instances'])):
        count=count+len(data['Reservations'][i]['Instances'])
        for k in range(len(data['Reservations'][i]['Instances'][j]['Tags'])):
            if data['Reservations'][i]['Instances'][j]['Tags'][k]['Key']=='Name':
                ins_name=data['Reservations'][i]['Instances'][j]['Tags'][k]['Value'].strip()
                ins_type=data['Reservations'][i]['Instances'][j]['InstanceType']
                state=data['Reservations'][i]['Instances'][j]['State']['Name']
                sh1.cell(row=count+1,column=1,value=ins_name)
                sh1.cell(row=count+1,column=2,value=ins_type)
                sh1.cell(row=count+1,column=3,value=state)
wb.save("ins_info1.xlsx")
wb.close()




for i in range(len(data['Reservations'])):
    for j in range(len(data['Reservations'][i]['Instances'])):
        for k in range(len(data['Reservations'][i]['Instances'][j]['Tags'])):
            if data['Reservations'][i]['Instances'][j]['Tags'][k]['Key']=='Name':
                print(data['Reservations'][i]['Instances'][j]['Tags'][k]['Value'].strip(), (data['Reservations'][i]['Instances'][j]['InstanceType']))



for i in range(len(data['Reservations'])):
    for j in range(len(data['Reservations'][i]['Instances'])):
        for k in range(len(data['Reservations'][i]['Instances'][j]['Tags'])):
            if data['Reservations'][i]['Instances'][j]['Tags'][k]['Key']=='Name':
                print(data['Reservations'][i]['Instances'][j]['Tags'][k]['Value'].strip(), (data['Reservations'][i]['Instances'][j]['InstanceType']))


#Curser move for defined time intervel, shoud run on top on chrome(tested) or any other applications(untested)
''' https://www.geeksforgeeks.org/mouse-keyboard-automation-using-python/ '''
import time
import pyautogui

x=0
y=700
for i in range(1000):
    pyautogui.moveTo(x,y,duration=1)
    pyautogui.click(x,y)
    time.sleep(2)
    x,y=y,x

#Listening to keyboard infinite time, stops when perticular key is pressed

from pynput import keyboard

def on_press(key):
    try:
        print('alphanumeric key {0} pressed'.format(key.char))
    except AttributeError:
        print('special key {0} pressed'.format(key))

def on_release(key):
    print('{0} released'.format(key))
    if key == keyboard.Key.esc:
        # Stop listener
        return False

#Collect events until released
with keyboard.Listener(
        on_press=on_press,
        on_release=on_release) as listener:
    listener.join()
'''
import pythoncom, pyHook 

def uMad(event):
    return False

hm = pyHook.HookManager()
hm.MouseAll = uMad
hm.KeyAll = uMad
hm.HookMouse()
hm.HookKeyboard()
pythoncom.PumpMessages()
'''


#Work on status of remote servers

import boto3
import fabric3
from fabric.api import run,env

ec2=boto3.client('ec2')


# How to serialize a datetime object as JSON using Python?
''' Got this error when dumping json varialbe into a local file 
Solution:
     The solution is quite simple. The json.dumps method can accept an optional parameter called default which is expected to be a function. Every time JSON tries to convert a value it does not know how to convert it will call the function we passed to it. The function will receive the object in question, and it is expected to return the JSON representation of the object.

     In the function we just call the __str__ method of the datetime object that will return a string     representation of the value. This is what we return.
     
     While the condition we have in the function is not required, if we expect to have other types of     objects in our data structure that need special treatment, we can make sure our function handles    them too. As dealing with each object will probably be different we check if the current object is     one we know to handle and do that inside the if statement. 
'''

import json
import datetime

def con(o):
    if isinstance(o,datetime.datetime):
        return o.__str__()
with open('feedsData.json','w') as outfile:
    print(json.dump(res,outfile,default=con)) #res is json response variable from AWS query
